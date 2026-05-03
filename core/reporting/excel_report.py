"""Excel report generator with branding."""
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


class ExcelReportGenerator:
    BRAND_BLUE = "2563EB"
    BRAND_NAVY = "0D2137"
    BRAND_TEAL = "0D9488"
    BRAND_RED = "DC2626"
    BRAND_AMBER = "D97706"
    BG_LIGHT = "FAFBFC"
    TEXT_PRIMARY = "111827"
    TEXT_SECONDARY = "6B7280"
    BORDER_COLOR = "E5E7EB"

    def __init__(self, tenant_name: str = "PDT Client"):
        self.tenant_name = tenant_name
        self._setup_styles()

    def _setup_styles(self):
        self.header_font = Font(name="Inter", bold=True, size=14, color=self.BRAND_NAVY)
        self.subheader_font = Font(name="Inter", bold=True, size=11, color=self.BRAND_NAVY)
        self.body_font = Font(name="Inter", size=10, color=self.TEXT_PRIMARY)
        self.mono_font = Font(name="JetBrains Mono", size=9, color=self.TEXT_PRIMARY)
        self.kpi_label_font = Font(name="Inter", size=9, color=self.TEXT_SECONDARY)
        self.kpi_value_font = Font(name="Inter", bold=True, size=16, color=self.BRAND_NAVY)

        self.navy_fill = PatternFill("solid", fgColor=self.BRAND_NAVY)
        self.blue_fill = PatternFill("solid", fgColor=self.BRAND_BLUE)
        self.teal_fill = PatternFill("solid", fgColor=self.BRAND_TEAL)
        self.red_fill = PatternFill("solid", fgColor=self.BRAND_RED)
        self.amber_fill = PatternFill("solid", fgColor=self.BRAND_AMBER)
        self.light_fill = PatternFill("solid", fgColor=self.BG_LIGHT)
        self.white_fill = PatternFill("solid", fgColor="FFFFFF")

        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color=self.BORDER_COLOR)
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set_column_width(self, ws, col: int, width: float):
        ws.column_dimensions[get_column_letter(col)].width = width

    def _apply_border_range(self, ws, min_row: int, max_row: int, min_col: int, max_col: int):
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = self.thin_border

    def generate(self, data: Dict[str, Any]) -> BytesIO:
        wb = Workbook()
        wb.remove(wb.active)

        self._build_executive_summary(wb, data)
        self._build_leakage_catalog(wb, data)
        self._build_action_plan(wb, data)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _build_executive_summary(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.create_sheet("Executive Summary")
        ws.sheet_properties.tabColor = self.BRAND_BLUE

        ws.merge_cells("A1:F1")
        ws["A1"] = f"B2B Pricing Diagnostic — Executive Summary"
        ws["A1"].font = Font(name="Inter", bold=True, size=18, color=self.BRAND_NAVY)
        ws["A1"].alignment = self.left_align
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Client: {self.tenant_name}  |  Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws["A2"].font = Font(name="Inter", size=10, color=self.TEXT_SECONDARY)
        ws["A2"].alignment = self.left_align

        kpis = data.get("kpis", {})
        row = 4
        ws.cell(row=row, column=1, value="KEY PERFORMANCE INDICATORS").font = self.subheader_font
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        kpi_cards = [
            ("Gross Margin %", f"{kpis.get('gross_margin_pct', 0):.1f}%", self.BRAND_TEAL),
            ("Pocket Price vs Invoice Δ", f"${kpis.get('pocket_vs_invoice_delta', 0):,.0f}", self.BRAND_BLUE),
            ("Leakages Found", kpis.get("leakages_found", 0), self.BRAND_RED),
            ("Total $ at Risk", f"${kpis.get('total_at_risk', 0):,.0f}", self.BRAND_AMBER),
        ]

        for i, (label, value, color) in enumerate(kpi_cards, start=1):
            col = i
            ws.cell(row=row, column=col, value=label).font = self.kpi_label_font
            ws.cell(row=row, column=col).alignment = self.center_align
            row += 1
            ws.cell(row=row, column=col, value=value).font = Font(
                name="Inter", bold=True, size=14, color=color
            )
            ws.cell(row=row, column=col).alignment = self.center_align
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="POCKET PRICE WATERFALL").font = self.subheader_font
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        waterfall = data.get("waterfall", [])
        headers = ["Stage", "Amount ($)", "% of List", "Δ from Prior"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=h).font = self.subheader_font
            ws.cell(row=row, column=col).fill = self.navy_fill
            ws.cell(row=row, column=col).font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
            ws.cell(row=row, column=col).alignment = self.center_align
        row += 1

        for item in waterfall:
            for col, val in enumerate([
                item.get("stage", ""),
                f"${item.get('amount', 0):,.2f}",
                f"{item.get('pct_of_list', 0):.1f}%",
                f"${item.get('delta', 0):,.2f}",
            ], start=1):
                ws.cell(row=row, column=col, value=val).font = self.body_font
                ws.cell(row=row, column=col).alignment = self.center_align
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="TOP 10 OPPORTUNITIES").font = self.subheader_font
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        opportunities = data.get("top_opportunities", [])
        for col, h in enumerate(["Rank", "Opportunity", "Category", "$ Impact", "Ease", "Priority"], start=1):
            ws.cell(row=row, column=col, value=h).font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
            ws.cell(row=row, column=col).fill = self.navy_fill
            ws.cell(row=row, column=col).alignment = self.center_align
        row += 1

        for rank, opp in enumerate(opportunities[:10], start=1):
            for col, val in enumerate([
                rank,
                opp.get("name", ""),
                opp.get("category", ""),
                f"${opp.get('impact', 0):,.0f}",
                opp.get("ease", 5),
                f"{opp.get('priority_score', 0):.1f}",
            ], start=1):
                ws.cell(row=row, column=col, value=val).font = self.body_font
                ws.cell(row=row, column=col).alignment = self.center_align
            row += 1

        for col in range(1, 7):
            self._set_column_width(ws, col, 22 if col == 2 else 15)

    def _build_leakage_catalog(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.create_sheet("Leakage Catalog")
        ws.sheet_properties.tabColor = self.BRAND_RED

        ws.merge_cells("A1:G1")
        ws["A1"] = "Leakage Catalog — All Rules"
        ws["A1"].font = Font(name="Inter", bold=True, size=16, color=self.BRAND_NAVY)

        row = 3
        for col, h in enumerate(["Rule ID", "Name", "Category", "Severity", "# Findings", "$ Impact", "Avg Confidence"], start=1):
            ws.cell(row=row, column=col, value=h).font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
            ws.cell(row=row, column=col).fill = self.navy_fill
            ws.cell(row=row, column=col).alignment = self.center_align
        row += 1

        rules = data.get("rule_summary", [])
        for rule in rules:
            for col, val in enumerate([
                rule.get("rule_id", ""),
                rule.get("name", ""),
                rule.get("category", ""),
                rule.get("severity", ""),
                rule.get("finding_count", 0),
                f"${rule.get('total_impact', 0):,.0f}",
                f"{rule.get('avg_confidence', 0):.2f}",
            ], start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = self.body_font
                cell.alignment = self.center_align
                if val == "HIGH":
                    cell.fill = self.red_fill
                    cell.font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
                elif val == "MEDIUM":
                    cell.fill = self.amber_fill
                    cell.font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
            row += 1

        for col in range(1, 8):
            self._set_column_width(ws, col, 18 if col == 2 else 14)

    def _build_action_plan(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.create_sheet("Action Plan")
        ws.sheet_properties.tabColor = self.BRAND_TEAL

        ws.merge_cells("A1:F1")
        ws["A1"] = "Prioritized Action Plan"
        ws["A1"].font = Font(name="Inter", bold=True, size=16, color=self.BRAND_NAVY)

        row = 3
        for col, h in enumerate(["Priority Rank", "Action Item", "Owner", "Effort", "Expected Return", "Status"], start=1):
            ws.cell(row=row, column=col, value=h).font = Font(name="Inter", bold=True, size=10, color="FFFFFF")
            ws.cell(row=row, column=col).fill = self.navy_fill
            ws.cell(row=row, column=col).alignment = self.center_align
        row += 1

        actions = data.get("action_plan", [])
        for rank, action in enumerate(actions, start=1):
            for col, val in enumerate([
                rank,
                action.get("item", ""),
                action.get("owner", "TBD"),
                action.get("effort", "Medium"),
                f"${action.get('expected_return', 0):,.0f}",
                "Open",
            ], start=1):
                ws.cell(row=row, column=col, value=val).font = self.body_font
                ws.cell(row=row, column=col).alignment = self.center_align
            row += 1

        for col in range(1, 7):
            self._set_column_width(ws, col, 28 if col == 2 else 14)
